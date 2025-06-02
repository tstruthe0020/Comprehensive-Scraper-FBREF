#!/usr/bin/env python3
"""
Excel Integrator - Specialized module for populating Excel files with FBref data
Works with the specific Excel structure: FBREF_Matches_[Seasons].xlsx
"""

import asyncio
import logging
from pathlib import Path
from typing import Dict, Any, List
import openpyxl
from openpyxl import load_workbook
import re

from fbref_batch_scraper import FBrefBatchScraper
from data_processor import DataProcessor
from config import Config

logger = logging.getLogger(__name__)

class ExcelIntegrator:
    def __init__(self, config: Config):
        self.config = config
        self.scraper = FBrefBatchScraper(config)
        self.data_processor = DataProcessor()
        
        # Excel cell mapping based on your structure
        self.match_metadata_mapping = {
            'season': (1, 2),
            'match_number': (2, 2),
            'match_url': (3, 2),
            'home_team': (4, 2),
            'away_team': (5, 2),
            'date': (6, 2),
            'competition': (7, 2),
            'source_url': (8, 2)
        }
        
        self.match_stats_mapping = {
            'home_goals': (12, 2),
            'away_goals': (13, 2),
            'final_score': (14, 2),
            'attendance': (15, 2),
            'referee': (16, 2),
            'stadium': (17, 2)
        }
        
        self.home_team_stats_mapping = {
            'possession': (22, 2),
            'total_shots': (23, 2),
            'shots_on_target': (24, 2),
            'corners': (25, 2),
            'fouls': (26, 2),
            'yellow_cards': (27, 2),
            'red_cards': (28, 2)
        }
        
        self.away_team_stats_mapping = {
            'possession': (32, 2),
            'total_shots': (33, 2),
            'shots_on_target': (34, 2),
            'corners': (35, 2),
            'fouls': (36, 2),
            'yellow_cards': (37, 2),
            'red_cards': (38, 2)
        }
        
        self.player_stats_columns = {
            'player_name': 1,
            'team': 2,
            'position': 3,
            'minutes': 4,
            'goals': 5,
            'assists': 6,
            'shots': 7,
            'passes': 8,
            'tackles': 9,
            'cards': 10
        }

    async def populate_excel_file(self, excel_file_path: str) -> Dict[str, Any]:
        """
        Main function to populate an Excel file with comprehensive FBref data
        """
        logger.info(f"Starting Excel population for: {excel_file_path}")
        
        try:
            # Load the Excel workbook
            wb = load_workbook(excel_file_path)
            
            # Get all match sheet names
            match_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith("Match_")]
            
            logger.info(f"Found {len(match_sheets)} match sheets to populate")
            
            # Setup browser for scraping
            if not await self.scraper.setup_browser():
                raise Exception("Failed to setup browser")
            
            results = {
                'total_matches': len(match_sheets),
                'successful_matches': 0,
                'failed_matches': 0,
                'errors': []
            }
            
            # Process each match sheet
            for i, sheet_name in enumerate(match_sheets):
                try:
                    logger.info(f"Processing sheet {i+1}/{len(match_sheets)}: {sheet_name}")
                    
                    # Get the worksheet
                    ws = wb[sheet_name]
                    
                    # Extract match URL from the sheet
                    match_url = ws.cell(row=3, column=2).value
                    
                    if not match_url:
                        logger.warning(f"No match URL found in sheet: {sheet_name}")
                        results['failed_matches'] += 1
                        results['errors'].append(f"No URL in {sheet_name}")
                        continue
                    
                    logger.info(f"Scraping: {match_url}")
                    
                    # Extract comprehensive data
                    season = ws.cell(row=1, column=2).value or "2024-25"
                    comprehensive_data = await self.scraper.extract_all_match_data(match_url, season)
                    
                    if not comprehensive_data:
                        logger.warning(f"No data extracted for: {match_url}")
                        results['failed_matches'] += 1
                        results['errors'].append(f"No data extracted from {match_url}")
                        continue
                    
                    # Process the data
                    processed_data = self.data_processor.process_comprehensive_data(comprehensive_data)
                    
                    # Populate the Excel sheet
                    self.populate_match_sheet(ws, processed_data, comprehensive_data)
                    
                    results['successful_matches'] += 1
                    logger.info(f"Successfully populated: {sheet_name}")
                    
                    # Rate limiting
                    if i < len(match_sheets) - 1:
                        await asyncio.sleep(self.config.RATE_LIMIT_DELAY)
                    
                except Exception as e:
                    logger.error(f"Error processing {sheet_name}: {e}")
                    results['failed_matches'] += 1
                    results['errors'].append(f"Error in {sheet_name}: {str(e)}")
            
            # Save the updated workbook
            wb.save(excel_file_path)
            logger.info(f"Excel file updated and saved: {excel_file_path}")
            
            # Update summary sheet if it exists
            if 'Summary' in wb.sheetnames:
                self.update_summary_sheet(wb['Summary'], results)
                wb.save(excel_file_path)
            
            return results
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            raise
        
        finally:
            await self.scraper.cleanup()

    def populate_match_sheet(self, worksheet, processed_data: Dict[str, Any], raw_data: Dict[str, Any]):
        """
        Populate a single match sheet with scraped data
        """
        try:
            match_info = processed_data.get('match_info', {})
            team_summary = processed_data.get('team_summary', [])
            player_stats = processed_data.get('player_stats', [])
            
            # 1. Populate Match Statistics (rows 12-17)
            self.populate_match_statistics(worksheet, match_info, raw_data)
            
            # 2. Populate Team Statistics (home: rows 22-28, away: rows 32-38)
            self.populate_team_statistics(worksheet, team_summary, match_info)
            
            # 3. Populate Player Statistics (starting from row 42)
            self.populate_player_statistics(worksheet, player_stats)
            
        except Exception as e:
            logger.error(f"Error populating match sheet: {e}")

    def populate_match_statistics(self, worksheet, match_info: Dict[str, Any], raw_data: Dict[str, Any]):
        """
        Populate match statistics section (rows 12-17)
        """
        try:
            # Extract scores from match info or raw data
            home_score, away_score = self.extract_scores(match_info, raw_data)
            
            # Populate match stats
            worksheet.cell(row=12, column=2, value=home_score)  # Goals (Home)
            worksheet.cell(row=13, column=2, value=away_score)  # Goals (Away)
            worksheet.cell(row=14, column=2, value=f"{home_score}-{away_score}")  # Final Score
            worksheet.cell(row=15, column=2, value=match_info.get('attendance', ''))  # Attendance
            worksheet.cell(row=16, column=2, value=match_info.get('referee', ''))  # Referee
            worksheet.cell(row=17, column=2, value=match_info.get('stadium', ''))  # Stadium
            
        except Exception as e:
            logger.error(f"Error populating match statistics: {e}")

    def populate_team_statistics(self, worksheet, team_summary: List[Dict[str, Any]], match_info: Dict[str, Any]):
        """
        Populate team statistics sections (home: rows 22-28, away: rows 32-38)
        """
        try:
            home_team = match_info.get('home_team', '').lower()
            away_team = match_info.get('away_team', '').lower()
            
            home_stats = {}
            away_stats = {}
            
            # Find stats for each team
            for team_stat in team_summary:
                team_name = team_stat.get('team', '').lower()
                
                if home_team and home_team in team_name:
                    home_stats = team_stat
                elif away_team and away_team in team_name:
                    away_stats = team_stat
                elif not home_stats:  # First team found
                    home_stats = team_stat
                elif not away_stats:  # Second team found
                    away_stats = team_stat
            
            # Populate home team stats (rows 22-28)
            self.populate_team_stats_section(worksheet, home_stats, self.home_team_stats_mapping)
            
            # Populate away team stats (rows 32-38)  
            self.populate_team_stats_section(worksheet, away_stats, self.away_team_stats_mapping)
            
        except Exception as e:
            logger.error(f"Error populating team statistics: {e}")

    def populate_team_stats_section(self, worksheet, team_stats: Dict[str, Any], mapping: Dict[str, tuple]):
        """
        Helper function to populate a team stats section
        """
        try:
            for stat_key, (row, col) in mapping.items():
                value = ""
                
                # Map our data fields to Excel fields
                if stat_key == 'possession':
                    value = team_stats.get('possession', '')
                elif stat_key == 'total_shots':
                    value = team_stats.get('shots_total', '')
                elif stat_key == 'shots_on_target':
                    value = team_stats.get('shots_on_target', '')
                elif stat_key == 'corners':
                    value = team_stats.get('corners', '')
                elif stat_key == 'fouls':
                    value = team_stats.get('fouls', '')
                elif stat_key == 'yellow_cards':
                    value = team_stats.get('cards_yellow', '')
                elif stat_key == 'red_cards':
                    value = team_stats.get('cards_red', '')
                
                # Clean up the value (remove % signs, convert to numbers if needed)
                if value and isinstance(value, str):
                    value = value.replace('%', '')
                    if value.isdigit():
                        value = int(value)
                    elif self.is_float(value):
                        value = float(value)
                
                worksheet.cell(row=row, column=col, value=value)
                
        except Exception as e:
            logger.error(f"Error populating team stats section: {e}")

    def populate_player_statistics(self, worksheet, player_stats: List[Dict[str, Any]]):
        """
        Populate player statistics section (starting from row 42)
        """
        try:
            # Clear existing player data first
            for row in range(42, 200):  # Clear up to row 200
                for col in range(1, 11):
                    worksheet.cell(row=row, column=col, value="")
            
            # Populate player data
            for i, player in enumerate(player_stats):
                row = 42 + i
                
                # Populate each column
                worksheet.cell(row=row, column=1, value=player.get('player_name', ''))  # Player Name
                worksheet.cell(row=row, column=2, value=player.get('team', ''))  # Team
                worksheet.cell(row=row, column=3, value=player.get('position', ''))  # Position
                worksheet.cell(row=row, column=4, value=self.clean_numeric_value(player.get('minutes', '')))  # Minutes
                worksheet.cell(row=row, column=5, value=self.clean_numeric_value(player.get('goals', '')))  # Goals
                worksheet.cell(row=row, column=6, value=self.clean_numeric_value(player.get('assists', '')))  # Assists
                worksheet.cell(row=row, column=7, value=self.clean_numeric_value(player.get('shots_total', '')))  # Shots
                worksheet.cell(row=row, column=8, value=self.clean_numeric_value(player.get('passes_completed', '')))  # Passes
                worksheet.cell(row=row, column=9, value=self.clean_numeric_value(player.get('tackles_won', '')))  # Tackles
                worksheet.cell(row=row, column=10, value=self.clean_numeric_value(player.get('cards_yellow', '')))  # Cards
                
        except Exception as e:
            logger.error(f"Error populating player statistics: {e}")

    def update_summary_sheet(self, summary_worksheet, results: Dict[str, Any]):
        """
        Update the summary sheet with scraping results
        """
        try:
            # Find the last row with data
            last_row = summary_worksheet.max_row
            
            # Add a summary row at the bottom
            summary_row = last_row + 2
            summary_worksheet.cell(row=summary_row, column=1, value="SCRAPING SUMMARY")
            summary_worksheet.cell(row=summary_row + 1, column=1, value="Total Matches:")
            summary_worksheet.cell(row=summary_row + 1, column=2, value=results['total_matches'])
            summary_worksheet.cell(row=summary_row + 2, column=1, value="Successful:")
            summary_worksheet.cell(row=summary_row + 2, column=2, value=results['successful_matches'])
            summary_worksheet.cell(row=summary_row + 3, column=1, value="Failed:")
            summary_worksheet.cell(row=summary_row + 3, column=2, value=results['failed_matches'])
            
        except Exception as e:
            logger.error(f"Error updating summary sheet: {e}")

    def extract_scores(self, match_info: Dict[str, Any], raw_data: Dict[str, Any]) -> tuple:
        """
        Extract home and away scores from available data
        """
        try:
            # Try to get scores from page title
            page_title = match_info.get('page_title', '')
            
            # Look for score pattern in title like "Arsenal 2-1 Chelsea"
            score_pattern = r'(\d+)[^\d]*(\d+)'
            match = re.search(score_pattern, page_title)
            
            if match:
                return int(match.group(1)), int(match.group(2))
            
            # Try to extract from raw data scorebox
            basic_info = raw_data.get('basic_info', {})
            scorebox_data = basic_info.get('scorebox_data', {})
            scorebox_text = scorebox_data.get('text_content', '')
            
            # Look for score in scorebox text
            match = re.search(score_pattern, scorebox_text)
            if match:
                return int(match.group(1)), int(match.group(2))
            
            return 0, 0
            
        except Exception as e:
            logger.error(f"Error extracting scores: {e}")
            return 0, 0

    def clean_numeric_value(self, value) -> int:
        """
        Clean and convert a value to integer
        """
        try:
            if not value:
                return 0
            
            if isinstance(value, (int, float)):
                return int(value)
            
            # Remove non-numeric characters except decimal point
            cleaned = re.sub(r'[^\d.]', '', str(value))
            
            if cleaned:
                return int(float(cleaned))
            
            return 0
            
        except:
            return 0

    def is_float(self, value: str) -> bool:
        """
        Check if a string can be converted to float
        """
        try:
            float(value)
            return True
        except:
            return False

    async def populate_from_summary_sheet(self, excel_file_path: str) -> Dict[str, Any]:
        """
        Alternative method: Extract URLs from Summary sheet and populate individual sheets
        """
        logger.info(f"Populating Excel from Summary sheet: {excel_file_path}")
        
        try:
            wb = load_workbook(excel_file_path)
            
            if 'Summary' not in wb.sheetnames:
                raise Exception("No Summary sheet found in Excel file")
            
            summary_ws = wb['Summary']
            
            # Extract match URLs from Summary sheet
            urls_and_sheets = []
            
            for row in range(2, summary_ws.max_row + 1):  # Skip header row
                match_url = summary_ws.cell(row=row, column=2).value  # Assuming URL is in column B
                sheet_name = summary_ws.cell(row=row, column=12).value  # Assuming sheet name is in column L
                
                if match_url and sheet_name:
                    urls_and_sheets.append((match_url, sheet_name))
            
            logger.info(f"Found {len(urls_and_sheets)} matches to process from Summary sheet")
            
            # Setup browser
            if not await self.scraper.setup_browser():
                raise Exception("Failed to setup browser")
            
            results = {
                'total_matches': len(urls_and_sheets),
                'successful_matches': 0,
                'failed_matches': 0,
                'errors': []
            }
            
            # Process each match
            for i, (match_url, sheet_name) in enumerate(urls_and_sheets):
                try:
                    if sheet_name not in wb.sheetnames:
                        logger.warning(f"Sheet {sheet_name} not found in workbook")
                        continue
                    
                    ws = wb[sheet_name]
                    
                    logger.info(f"Processing {i+1}/{len(urls_and_sheets)}: {sheet_name}")
                    
                    # Extract comprehensive data
                    comprehensive_data = await self.scraper.extract_all_match_data(match_url, "2024-25")
                    
                    if comprehensive_data:
                        processed_data = self.data_processor.process_comprehensive_data(comprehensive_data)
                        self.populate_match_sheet(ws, processed_data, comprehensive_data)
                        results['successful_matches'] += 1
                    else:
                        results['failed_matches'] += 1
                        results['errors'].append(f"No data for {match_url}")
                    
                    # Rate limiting
                    if i < len(urls_and_sheets) - 1:
                        await asyncio.sleep(self.config.RATE_LIMIT_DELAY)
                    
                except Exception as e:
                    logger.error(f"Error processing {sheet_name}: {e}")
                    results['failed_matches'] += 1
                    results['errors'].append(f"Error in {sheet_name}: {str(e)}")
            
            # Save workbook
            wb.save(excel_file_path)
            self.update_summary_sheet(wb['Summary'], results)
            wb.save(excel_file_path)
            
            return results
            
        finally:
            await self.scraper.cleanup()

# CLI Interface for Excel Integration
async def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel FBref Data Populator')
    parser.add_argument('--excel', required=True, help='Path to Excel file')
    parser.add_argument('--method', choices=['individual', 'summary'], default='individual', 
                       help='Population method: individual sheets or from summary sheet')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Initialize integrator
    config = Config()
    config.RATE_LIMIT_DELAY = 3  # Be conservative with Excel processing
    
    integrator = ExcelIntegrator(config)
    
    try:
        if args.method == 'summary':
            results = await integrator.populate_from_summary_sheet(args.excel)
        else:
            results = await integrator.populate_excel_file(args.excel)
        
        # Print results
        print(f"\n{'='*50}")
        print(f"EXCEL POPULATION COMPLETED")
        print(f"{'='*50}")
        print(f"Total Matches: {results['total_matches']}")
        print(f"Successful: {results['successful_matches']}")
        print(f"Failed: {results['failed_matches']}")
        print(f"Success Rate: {(results['successful_matches']/results['total_matches']*100):.1f}%")
        
        if results['errors']:
            print(f"\nErrors:")
            for error in results['errors'][:5]:  # Show first 5 errors
                print(f"  - {error}")
        
        print(f"\nExcel file updated: {args.excel}")
        
    except Exception as e:
        logger.error(f"Excel population failed: {e}")
        print(f"Error: {e}")

if __name__ == "__main__":
    asyncio.run(main())
