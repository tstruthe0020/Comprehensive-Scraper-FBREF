#!/usr/bin/env python3
"""
Create a test Excel file matching the exact structure described
"""

import openpyxl
from openpyxl import Workbook

def create_test_excel_matching_structure():
    """Create Excel file matching the exact structure described"""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Summary sheet headers (as described)
    headers = [
        "Season", "Match_Report_URL", "Match_Number", "Home_Team", "Away_Team", 
        "Date", "Score", "Home_Goals", "Away_Goals", "Competition", "Venue", "Sheet_Name"
    ]
    
    for col, header in enumerate(headers, 1):
        summary_ws.cell(row=1, column=col, value=header)
    
    # Add sample data matching the structure
    sample_matches = [
        [
            "2023-2024", 
            "https://fbref.com/en/matches/cc5b4244/Manchester-United-Fulham-August-16-2024-Premier-League",
            1, "Manchester United", "Fulham", "August-16-2024", "", "", "", "Premier League", 
            "Old Trafford", "Match_001_Manchester_vs_Fulham"
        ],
        [
            "2023-2024",
            "https://fbref.com/en/matches/8b1e4321/Arsenal-Wolverhampton-Wanderers-August-17-2024-Premier-League", 
            2, "Arsenal", "Wolves", "August-17-2024", "", "", "", "Premier League",
            "Emirates Stadium", "Match_002_Arsenal_vs_Wolves"
        ],
        [
            "2023-2024",
            "https://fbref.com/en/matches/9c2f5432/Brighton-Hove-Albion-Everton-August-17-2024-Premier-League",
            3, "Brighton", "Everton", "August-17-2024", "", "", "", "Premier League", 
            "Amex Stadium", "Match_003_Brighton_vs_Everton"
        ]
    ]
    
    for row, data in enumerate(sample_matches, 2):
        for col, value in enumerate(data, 1):
            summary_ws.cell(row=row, column=col, value=value)
    
    # Create individual match sheets with EXACT structure described
    for match_data in sample_matches:
        season, match_url, match_num, home_team, away_team, date, _, _, _, comp, venue, sheet_name = match_data
        
        ws = wb.create_sheet(sheet_name)
        
        # EXACT STRUCTURE AS DESCRIBED:
        
        # A. Match Metadata Section (Rows 1-10)
        metadata_fields = [
            ("Season", season),
            ("Match Number", match_num), 
            ("Match Report URL", match_url),
            ("Home Team", home_team),
            ("Away Team", away_team),
            ("Date", date),
            ("Competition", comp),
            ("Source URL", "https://fbref.com/en/comps/9/schedule/...")
        ]
        
        for row, (field, value) in enumerate(metadata_fields, 1):
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value=value)
        
        # B. Match Statistics Section (Rows 11-20)
        ws.cell(row=11, column=1, value="MATCH STATISTICS")
        match_stats_fields = [
            "Goals (Home)",
            "Goals (Away)", 
            "Final Score",
            "Attendance",
            "Referee",
            "Stadium"
        ]
        
        for row, field in enumerate(match_stats_fields, 12):
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value="[EMPTY]")  # As described
        
        # C. Home Team Stats Section (Rows 21-30)
        ws.cell(row=21, column=1, value="HOME TEAM STATS")
        home_stats_fields = [
            "Possession (%)",
            "Total Shots",
            "Shots on Target", 
            "Corners",
            "Fouls",
            "Yellow Cards",
            "Red Cards"
        ]
        
        for row, field in enumerate(home_stats_fields, 22):
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value="[EMPTY]")
        
        # D. Away Team Stats Section (Rows 31-40)
        ws.cell(row=31, column=1, value="AWAY TEAM STATS")
        away_stats_fields = [
            "Possession (%)",
            "Total Shots", 
            "Shots on Target",
            "Corners",
            "Fouls",
            "Yellow Cards",
            "Red Cards"
        ]
        
        for row, field in enumerate(away_stats_fields, 32):
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value="[EMPTY]")
        
        # E. Player Statistics Section (Rows 41+)
        ws.cell(row=41, column=1, value="PLAYER STATISTICS")
        
        # Player stats headers (Row 42)
        player_headers = [
            "Player Name", "Team", "Position", "Minutes", "Goals", 
            "Assists", "Shots", "Passes", "Tackles", "Cards"
        ]
        
        for col, header in enumerate(player_headers, 1):
            ws.cell(row=42, column=col, value=header)
        
        # Add empty player rows (as described)
        for row in range(43, 48):  # Add 5 empty player rows
            for col in range(1, 11):
                ws.cell(row=row, column=col, value="[EMPTY]")
    
    # Save the test file
    test_file = "TEST_FBREF_Structure.xlsx"
    wb.save(test_file)
    print(f"✅ Created test Excel file: {test_file}")
    print("📋 This matches your exact structure:")
    print("   - Summary sheet with match index")
    print("   - Individual match sheets (Match_001_, Match_002_, etc.)")
    print("   - Metadata in rows 1-8")
    print("   - Match stats in rows 12-17")
    print("   - Home team stats in rows 22-28")
    print("   - Away team stats in rows 32-38") 
    print("   - Player stats starting row 42")
    
    return test_file

if __name__ == "__main__":
    create_test_excel_matching_structure()