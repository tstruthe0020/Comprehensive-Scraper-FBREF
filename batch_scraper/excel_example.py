#!/usr/bin/env python3
"""
Example script showing how to use the Excel Integrator
"""

import asyncio
import logging
from pathlib import Path
from excel_integrator import ExcelIntegrator
from config import Config

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

async def demo_excel_population():
    """
    Demonstrate Excel population with a sample file
    """
    
    print("🚀 FBref Excel Integrator Demo")
    print("=" * 50)
    
    # Configuration
    config = Config()
    config.HEADLESS = True  # Set to False to see browser actions
    config.RATE_LIMIT_DELAY = 3  # Conservative rate limiting
    
    # Initialize integrator
    integrator = ExcelIntegrator(config)
    
    # Example Excel file path (you would use your actual file)
    excel_file = "FBREF_Matches_2024-25.xlsx"
    
    if not Path(excel_file).exists():
        print(f"❌ Excel file not found: {excel_file}")
        print("Please create an Excel file with the structure described in the documentation")
        return
    
    try:
        print(f"📊 Processing Excel file: {excel_file}")
        print("This will:")
        print("  ✅ Read match URLs from individual sheets")
        print("  ✅ Scrape comprehensive data from FBref")
        print("  ✅ Populate all match statistics")
        print("  ✅ Fill team statistics (home & away)")
        print("  ✅ Add player statistics")
        print("  ✅ Update match metadata")
        
        # Populate the Excel file
        results = await integrator.populate_excel_file(excel_file)
        
        # Display results
        print(f"\n{'🎉 RESULTS' : ^50}")
        print("=" * 50)
        print(f"📈 Total Matches Processed: {results['total_matches']}")
        print(f"✅ Successful Extractions: {results['successful_matches']}")
        print(f"❌ Failed Extractions: {results['failed_matches']}")
        
        if results['total_matches'] > 0:
            success_rate = (results['successful_matches'] / results['total_matches']) * 100
            print(f"📊 Success Rate: {success_rate:.1f}%")
        
        if results['errors']:
            print(f"\n⚠️  Errors encountered:")
            for i, error in enumerate(results['errors'][:3], 1):
                print(f"   {i}. {error}")
            if len(results['errors']) > 3:
                print(f"   ... and {len(results['errors']) - 3} more errors")
        
        print(f"\n💾 Excel file updated: {excel_file}")
        print("📋 Data populated in your existing sheet structure!")
        
    except Exception as e:
        print(f"❌ Error during Excel population: {e}")
        logger.error(f"Demo failed: {e}")

def create_sample_excel_structure():
    """
    Create a sample Excel file with the proper structure for testing
    """
    import openpyxl
    from openpyxl import Workbook
    
    print("📝 Creating sample Excel structure...")
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Summary sheet headers
    headers = ["Season", "Match_Report_URL", "Match_Number", "Home_Team", "Away_Team", 
              "Date", "Score", "Home_Goals", "Away_Goals", "Competition", "Venue", "Sheet_Name"]
    
    for col, header in enumerate(headers, 1):
        summary_ws.cell(row=1, column=col, value=header)
    
    # Add sample data
    sample_data = [
        ["2024-25", "https://fbref.com/en/matches/d14fd5fb/Arsenal-Tottenham-September-15-2024-Premier-League", 
         1, "Arsenal", "Tottenham", "2024-09-15", "", "", "", "Premier League", "Emirates Stadium", "Match_001_Arsenal_vs_Tottenham"],
        ["2024-25", "https://fbref.com/en/matches/a7f2d9e1/Liverpool-Chelsea-January-26-2025-Premier-League", 
         2, "Liverpool", "Chelsea", "2025-01-26", "", "", "", "Premier League", "Anfield", "Match_002_Liverpool_vs_Chelsea"]
    ]
    
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            summary_ws.cell(row=row, column=col, value=value)
    
    # Create individual match sheets
    for i, (season, url, match_num, home, away, date, _, _, _, comp, venue, sheet_name) in enumerate(sample_data):
        ws = wb.create_sheet(sheet_name)
        
        # Match Metadata Section (Rows 1-10)
        ws.cell(row=1, column=1, value="Season")
        ws.cell(row=1, column=2, value=season)
        ws.cell(row=2, column=1, value="Match Number")
        ws.cell(row=2, column=2, value=match_num)
        ws.cell(row=3, column=1, value="Match Report URL")
        ws.cell(row=3, column=2, value=url)
        ws.cell(row=4, column=1, value="Home Team")
        ws.cell(row=4, column=2, value=home)
        ws.cell(row=5, column=1, value="Away Team")
        ws.cell(row=5, column=2, value=away)
        ws.cell(row=6, column=1, value="Date")
        ws.cell(row=6, column=2, value=date)
        ws.cell(row=7, column=1, value="Competition")
        ws.cell(row=7, column=2, value=comp)
        ws.cell(row=8, column=1, value="Source URL")
        ws.cell(row=8, column=2, value="")
        
        # Match Statistics Section (Rows 11-20)
        ws.cell(row=11, column=1, value="MATCH STATISTICS")
        ws.cell(row=12, column=1, value="Goals (Home)")
        ws.cell(row=13, column=1, value="Goals (Away)")
        ws.cell(row=14, column=1, value="Final Score")
        ws.cell(row=15, column=1, value="Attendance")
        ws.cell(row=16, column=1, value="Referee")
        ws.cell(row=17, column=1, value="Stadium")
        
        # Home Team Stats Section (Rows 21-30)
        ws.cell(row=21, column=1, value="HOME TEAM STATS")
        ws.cell(row=22, column=1, value="Possession (%)")
        ws.cell(row=23, column=1, value="Total Shots")
        ws.cell(row=24, column=1, value="Shots on Target")
        ws.cell(row=25, column=1, value="Corners")
        ws.cell(row=26, column=1, value="Fouls")
        ws.cell(row=27, column=1, value="Yellow Cards")
        ws.cell(row=28, column=1, value="Red Cards")
        
        # Away Team Stats Section (Rows 31-40)
        ws.cell(row=31, column=1, value="AWAY TEAM STATS")
        ws.cell(row=32, column=1, value="Possession (%)")
        ws.cell(row=33, column=1, value="Total Shots")
        ws.cell(row=34, column=1, value="Shots on Target")
        ws.cell(row=35, column=1, value="Corners")
        ws.cell(row=36, column=1, value="Fouls")
        ws.cell(row=37, column=1, value="Yellow Cards")
        ws.cell(row=38, column=1, value="Red Cards")
        
        # Player Statistics Section (Rows 41+)
        ws.cell(row=41, column=1, value="PLAYER STATISTICS")
        player_headers = ["Player Name", "Team", "Position", "Minutes", "Goals", "Assists", "Shots", "Passes", "Tackles", "Cards"]
        for col, header in enumerate(player_headers, 1):
            ws.cell(row=42, column=col, value=header)
    
    # Save the sample file
    sample_file = "Sample_FBREF_Matches_2024-25.xlsx"
    wb.save(sample_file)
    print(f"✅ Sample Excel file created: {sample_file}")
    print("📋 You can use this file to test the Excel integrator")
    
    return sample_file

async def main():
    print("FBref Excel Integrator")
    print("=" * 30)
    print("1. Create sample Excel structure")
    print("2. Run demo population (requires existing Excel file)")
    print("3. Exit")
    
    choice = input("\nEnter your choice (1-3): ").strip()
    
    if choice == "1":
        create_sample_excel_structure()
    elif choice == "2":
        await demo_excel_population()
    elif choice == "3":
        print("Goodbye!")
    else:
        print("Invalid choice. Please run again.")

if __name__ == "__main__":
    asyncio.run(main())
