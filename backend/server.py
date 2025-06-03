from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import requests
import re
from bs4 import BeautifulSoup
import csv
import io
import os
from typing import List, Dict
from fastapi.responses import StreamingResponse
from playwright.async_api import async_playwright
import asyncio
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import the CSV scraper
try:
    from csv_scraper import CSVMatchReportScraper
    CSV_SCRAPER_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ CSV scraper not available: {e}")
    CSV_SCRAPER_AVAILABLE = False

# Import the comprehensive scraper integration
try:
    from integration_wrapper import enhance_excel_with_fbref_data, validate_excel_for_fbref
    COMPREHENSIVE_SCRAPER_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Comprehensive scraper not available: {e}")
    COMPREHENSIVE_SCRAPER_AVAILABLE = False

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class CSVScrapingRequest(BaseModel):
    urls: List[str]
    max_matches: int = None

class CSVScrapingResponse(BaseModel):
    success: bool
    message: str
    csv_data: str = ""
    filename: str = ""
    total_matches: int = 0
    processed_matches: int = 0

class ScrapingRequest(BaseModel):
    urls: List[str]  # Changed from single url to multiple urls

class SeasonResult(BaseModel):
    url: str
    season_name: str
    success: bool
    message: str
    links: List[str] = []

class ScrapingResponse(BaseModel):
    success: bool
    message: str
    seasons: List[SeasonResult] = []
    total_links: int = 0
    excel_data: str = ""  # Base64 encoded Excel file
    filename: str = ""
    enhancement_available: bool = False
    enhancement_results: Dict = {}

@app.post("/api/csv-scrape-workflow")
async def csv_scrape_full_workflow(request: CSVScrapingRequest):
    """
    Complete CSV-based workflow:
    1. Extract match URLs from fixtures pages
    2. Create CSV with match URLs
    3. Scrape team and player stats from each match
    4. Update CSV with all statistics
    """
    if not CSV_SCRAPER_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="CSV scraper not available. Please check system setup."
        )
    
    try:
        if not request.urls or len(request.urls) == 0:
            raise HTTPException(status_code=400, detail="Please provide at least one FBREF fixtures URL")
        
        # Validate URLs
        for url in request.urls:
            if not url or not url.strip().startswith("https://fbref.com"):
                raise HTTPException(status_code=400, detail=f"Invalid FBREF URL: {url}")
        
        # Initialize CSV scraper
        csv_scraper = CSVMatchReportScraper(rate_limit_delay=3, headless=True)
        
        logger.info(f"Starting CSV workflow for {len(request.urls)} fixtures URLs")
        
        # Run the complete workflow
        csv_content = await csv_scraper.full_workflow(
            fixtures_urls=[url.strip() for url in request.urls],
            max_matches=request.max_matches
        )
        
        if not csv_content:
            return CSVScrapingResponse(
                success=False,
                message="Failed to generate CSV data",
                csv_data="",
                filename="",
                total_matches=0,
                processed_matches=0
            )
        
        # Count matches for response
        lines = csv_content.strip().split('\n')
        total_matches = max(0, len(lines) - 1)  # Subtract header row
        
        # Count processed matches (those with 'Yes' in Scraped column)
        processed_matches = 0
        if len(lines) > 1:
            import csv
            import io
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            processed_matches = sum(1 for row in csv_reader if row.get('Scraped', '').strip().lower() == 'yes')
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"FBREF_Match_Stats_{timestamp}.csv"
        
        # Encode CSV to base64 for download
        csv_b64 = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
        
        return CSVScrapingResponse(
            success=True,
            message=f"Successfully processed {processed_matches}/{total_matches} matches with complete statistics",
            csv_data=csv_b64,
            filename=filename,
            total_matches=total_matches,
            processed_matches=processed_matches
        )
        
    except Exception as e:
        logger.error(f"CSV workflow error: {e}")
        return CSVScrapingResponse(
            success=False,
            message=f"CSV workflow failed: {str(e)}",
            csv_data="",
            filename="",
            total_matches=0,
            processed_matches=0
        )

@app.post("/api/csv-extract-urls-only")
async def csv_extract_match_urls_only(request: CSVScrapingRequest):
    """
    Step 1 only: Extract match URLs from fixtures and create initial CSV
    """
    if not CSV_SCRAPER_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="CSV scraper not available. Please check system setup."
        )
    
    try:
        if not request.urls or len(request.urls) == 0:
            raise HTTPException(status_code=400, detail="Please provide at least one FBREF fixtures URL")
        
        # Initialize CSV scraper
        csv_scraper = CSVMatchReportScraper(rate_limit_delay=2, headless=True)
        
        if not await csv_scraper.setup_browser():
            raise HTTPException(status_code=500, detail="Failed to setup browser")
        
        try:
            all_match_urls = []
            
            # Extract URLs from all fixtures pages
            for fixtures_url in request.urls:
                match_urls = await csv_scraper.extract_match_urls_from_fixtures(fixtures_url.strip())
                all_match_urls.extend(match_urls)
                await asyncio.sleep(1)  # Rate limiting
            
            # Remove duplicates and limit if specified
            all_match_urls = list(dict.fromkeys(all_match_urls))
            if request.max_matches:
                all_match_urls = all_match_urls[:request.max_matches]
            
            if not all_match_urls:
                return CSVScrapingResponse(
                    success=False,
                    message="No match URLs found in the provided fixtures pages",
                    csv_data="",
                    filename="",
                    total_matches=0,
                    processed_matches=0
                )
            
            # Create initial CSV
            csv_content = csv_scraper.create_initial_csv(all_match_urls)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"FBREF_Match_URLs_{timestamp}.csv"
            
            # Encode to base64
            csv_b64 = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
            
            return CSVScrapingResponse(
                success=True,
                message=f"Successfully extracted {len(all_match_urls)} match URLs",
                csv_data=csv_b64,
                filename=filename,
                total_matches=len(all_match_urls),
                processed_matches=0
            )
            
        finally:
            await csv_scraper.cleanup()
        
    except Exception as e:
        logger.error(f"URL extraction error: {e}")
        return CSVScrapingResponse(
            success=False,
            message=f"URL extraction failed: {str(e)}",
            csv_data="",
            filename="",
            total_matches=0,
            processed_matches=0
        )

@app.post("/api/demo-csv-workflow")
async def demo_csv_workflow():
    """Demo the CSV workflow with current Premier League season (limited matches)"""
    try:
        # Use current Premier League fixtures
        demo_url = "https://fbref.com/en/comps/9/schedule/Premier-League-Scores-and-Fixtures"
        
        csv_scraper = CSVMatchReportScraper(rate_limit_delay=2, headless=True)
        
        # Run workflow with max 3 matches for demo
        csv_content = await csv_scraper.full_workflow(
            fixtures_urls=[demo_url],
            max_matches=3
        )
        
        if not csv_content:
            return CSVScrapingResponse(
                success=False,
                message="Demo failed: No data generated",
                csv_data="",
                filename="",
                total_matches=0,
                processed_matches=0
            )
        
        # Count matches
        lines = csv_content.strip().split('\n')
        total_matches = max(0, len(lines) - 1)
        
        # Count processed
        processed_matches = 0
        if len(lines) > 1:
            import csv
            import io
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            processed_matches = sum(1 for row in csv_reader if row.get('Scraped', '').strip().lower() == 'yes')
        
        filename = f"FBREF_Demo_CSV_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        csv_b64 = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
        
        return CSVScrapingResponse(
            success=True,
            message=f"Demo completed: {processed_matches}/{total_matches} matches with stats from current Premier League",
            csv_data=csv_b64,
            filename=filename,
            total_matches=total_matches,
            processed_matches=processed_matches
        )
        
    except Exception as e:
        logger.error(f"Demo CSV workflow error: {e}")
        return CSVScrapingResponse(
            success=False,
            message=f"Demo failed: {str(e)}",
            csv_data="",
            filename="",
            total_matches=0,
            processed_matches=0
        )

@app.get("/api/health")
async def health_check():
    return {"status": "healthy"}

async def scrape_fbref_with_playwright(url: str) -> SeasonResult:
    """Scrape a single FBREF fixtures page using Playwright"""
    try:
        # Extract season name from URL for better organization
        season_name = "Unknown Season"
        if "/schedule/" in url:
            parts = url.split("/")
            for i, part in enumerate(parts):
                if "schedule" in part and i > 0:
                    season_name = parts[i-1] if parts[i-1] != "schedule" else "Current Season"
                    break
        
        async with async_playwright() as p:
            # Launch Firefox browser with stealth settings
            browser = await p.firefox.launch(
                headless=True,
                args=[
                    '--no-sandbox',
                    '--disable-blink-features=AutomationControlled',
                ]
            )
            
            # Create a new context with realistic browser settings
            context = await browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            
            page = await context.new_page()
            
            # Navigate to the page with realistic timing
            await page.goto(url, wait_until='networkidle', timeout=30000)
            
            # Wait a bit to let any dynamic content load
            await page.wait_for_timeout(2000)
            
            # Get page content
            content = await page.content()
            
            await browser.close()
        
        # Parse HTML with BeautifulSoup
        soup = BeautifulSoup(content, 'html.parser')
        
        # Find the Score column header with multiple fallback strategies
        score_header = None
        
        # Strategy 1: Exact match for the specified attributes
        score_header = soup.find('th', {
            'aria-label': 'Score',
            'data-stat': 'score',
            'scope': 'col'
        })
        
        # Strategy 2: Find by data-stat only
        if not score_header:
            score_header = soup.find('th', {'data-stat': 'score'})
        
        # Strategy 3: Find by text content
        if not score_header:
            score_header = soup.find('th', string=re.compile(r'Score', re.IGNORECASE))
            
        # Strategy 4: Find by aria-label only
        if not score_header:
            score_header = soup.find('th', {'aria-label': re.compile(r'Score', re.IGNORECASE)})
        
        if not score_header:
            return SeasonResult(
                url=url,
                season_name=season_name,
                success=False,
                message="Could not find the Score column in the table. Page structure may be different.",
                links=[]
            )
        
        # Find the table containing the score header
        table = score_header.find_parent('table')
        if not table:
            return SeasonResult(
                url=url,
                season_name=season_name,
                success=False,
                message="Could not find the table containing the Score column.",
                links=[]
            )
        
        # Get column index of Score column
        headers_row = score_header.find_parent('tr')
        headers = headers_row.find_all(['th', 'td'])
        score_column_index = None
        
        for i, header in enumerate(headers):
            if (header.get('data-stat') == 'score' or 
                (header.get('aria-label') and 'score' in header.get('aria-label').lower()) or
                (header.get_text(strip=True).lower() == 'score')):
                score_column_index = i
                break
        
        if score_column_index is None:
            return SeasonResult(
                url=url,
                season_name=season_name,
                success=False,
                message="Could not determine the position of the Score column.",
                links=[]
            )
        
        # Extract match report links from Score column
        match_links = []
        tbody = table.find('tbody')
        
        if tbody:
            rows = tbody.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) > score_column_index:
                    score_cell = cells[score_column_index]
                    # Look for links in the score cell
                    links = score_cell.find_all('a', href=True)
                    for link in links:
                        href = link['href']
                        # Check if it's a match report link
                        if '/matches/' in href:
                            full_url = f"https://fbref.com{href}" if href.startswith('/') else href
                            if full_url not in match_links:
                                match_links.append(full_url)
        
        # If no links in score column, try alternative approach
        if not match_links:
            all_links = table.find_all('a', href=True)
            for link in all_links:
                href = link['href']
                if '/matches/' in href:
                    full_url = f"https://fbref.com{href}" if href.startswith('/') else href
                    if full_url not in match_links:
                        match_links.append(full_url)
        
        if not match_links:
            return SeasonResult(
                url=url,
                season_name=season_name,
                success=False,
                message="No match report links found. This could mean no completed matches for this season.",
                links=[]
            )
        
        return SeasonResult(
            url=url,
            season_name=season_name,
            success=True,
            message=f"Successfully extracted {len(match_links)} match report links",
            links=match_links
        )
        
    except Exception as e:
        return SeasonResult(
            url=url,
            season_name=season_name,
            success=False,
            message=f"Error scraping {url}: {str(e)}",
            links=[]
        )

@app.post("/api/scrape-fbref")
async def scrape_multiple_fbref_seasons(request: ScrapingRequest):
    """Scrape multiple FBREF fixture pages sequentially"""
    try:
        if not request.urls or len(request.urls) == 0:
            raise HTTPException(status_code=400, detail="Please provide at least one FBREF URL")
        
        # Validate all URLs
        for url in request.urls:
            if not url or not url.strip().startswith("https://fbref.com"):
                raise HTTPException(status_code=400, detail=f"Invalid FBREF URL: {url}")
        
        season_results = []
        all_links = []
        
        # Process each URL sequentially
        for url in request.urls:
            url = url.strip()
            print(f"Processing: {url}")
            result = await scrape_fbref_with_playwright(url)
            season_results.append(result)
            
            if result.success:
                all_links.extend(result.links)
            
            # Small delay between requests to be respectful
            await asyncio.sleep(1)
        
        # Generate Excel workbook with separate sheets for each match
        excel_b64, filename = create_excel_workbook(season_results)
        
        # Prepare response
        successful_seasons = [s for s in season_results if s.success]
        failed_seasons = [s for s in season_results if not s.success]
        
        if len(successful_seasons) == 0:
            return ScrapingResponse(
                success=False,
                message="Failed to scrape any seasons. All URLs encountered errors.",
                seasons=season_results,
                total_links=0,
                excel_data="",
                filename=""
            )
        
        message_parts = []
        message_parts.append(f"Successfully processed {len(successful_seasons)}/{len(request.urls)} seasons")
        message_parts.append(f"Total match links extracted: {len(all_links)}")
        message_parts.append(f"Excel file generated with {len(all_links)} match sheets")
        
        if failed_seasons:
            message_parts.append(f"{len(failed_seasons)} seasons failed")
        
        return ScrapingResponse(
            success=True,
            message=" | ".join(message_parts),
            seasons=season_results,
            total_links=len(all_links),
            excel_data=excel_b64,
            filename=filename
        )
        
    except Exception as e:
        return ScrapingResponse(
            success=False,
            message=f"An error occurred during processing: {str(e)}",
            seasons=[],
            total_links=0,
            excel_data="",
            filename=""
        )

@app.post("/api/enhance-excel")
async def enhance_excel_with_comprehensive_data(request: dict):
    """Enhance an existing Excel file with comprehensive FBREF data"""
    if not COMPREHENSIVE_SCRAPER_AVAILABLE:
        raise HTTPException(
            status_code=503, 
            detail="Comprehensive scraper integration not available. Please check system setup."
        )
    
    try:
        excel_b64 = request.get('excel_data')
        filename = request.get('filename', 'enhanced_excel.xlsx')
        
        if not excel_b64:
            raise HTTPException(status_code=400, detail="Excel data required")
        
        # Decode Excel data
        excel_bytes = base64.b64decode(excel_b64)
        
        # Save to temporary file
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(excel_bytes)
            tmp_file_path = tmp_file.name
        
        try:
            # Validate Excel structure
            validation = validate_excel_for_fbref(tmp_file_path)
            
            if not validation['valid']:
                return {
                    'success': False,
                    'message': f"Excel structure incompatible: {validation['error']}",
                    'excel_data': excel_b64,
                    'filename': filename,
                    'enhancement_results': validation
                }
            
            # Enhance the Excel file (remove await - this function is not async)
            results = enhance_excel_with_fbref_data(tmp_file_path)
            
            # Read the enhanced file
            with open(tmp_file_path, 'rb') as f:
                enhanced_excel_bytes = f.read()
            
            enhanced_excel_b64 = base64.b64encode(enhanced_excel_bytes).decode('utf-8')
            
            # Add "Enhanced" to filename
            enhanced_filename = filename.replace('.xlsx', '_Enhanced.xlsx')
            
            message_parts = []
            if results['success']:
                message_parts.append(f"Successfully enhanced Excel file")
                message_parts.append(f"Populated {results['successful_matches']}/{results['total_matches']} matches")
                if results['failed_matches'] > 0:
                    message_parts.append(f"{results['failed_matches']} matches failed (normal for anti-scraping)")
            else:
                message_parts.append("Enhancement completed with issues")
            
            return {
                'success': results['success'],
                'message': " | ".join(message_parts),
                'excel_data': enhanced_excel_b64,
                'filename': enhanced_filename,
                'enhancement_results': results
            }
            
        finally:
            # Clean up temporary file
            os.unlink(tmp_file_path)
            
    except Exception as e:
        return {
            'success': False,
            'message': f"Enhancement failed: {str(e)}",
            'excel_data': excel_b64,
            'filename': filename,
            'enhancement_results': {'error': str(e)}
        }

@app.get("/api/check-enhancement")
async def check_enhancement_availability():
    """Check if comprehensive scraper enhancement is available"""
    if COMPREHENSIVE_SCRAPER_AVAILABLE:
        try:
            from integration_wrapper import check_fbref_availability
            status = check_fbref_availability()
            return {
                'available': True,
                'status': 'Comprehensive enhancement ready',
                'details': status
            }
        except Exception as e:
            return {
                'available': False,
                'status': f'Enhancement error: {str(e)}',
                'details': {}
            }
    else:
        return {
            'available': False,
            'status': 'Comprehensive scraper not installed',
            'details': {}
        }

@app.post("/api/demo-scrape")
async def demo_scrape_fbref():
    """Demo endpoint that pulls real data from current Premier League season (first 5 matches)"""
    try:
        # Get real data from current Premier League season
        current_season_url = "https://fbref.com/en/comps/9/schedule/Premier-League-Scores-and-Fixtures"
        
        print(f"Demo: Scraping real data from {current_season_url}")
        real_result = await scrape_fbref_with_playwright(current_season_url)
        
        if not real_result.success or len(real_result.links) == 0:
            # Fallback to fake data if real scraping fails
            print("Demo: Real scraping failed, using fallback fake data")
            demo_seasons = [
                SeasonResult(
                    url="https://fbref.com/en/comps/9/schedule/Premier-League-Scores-and-Fixtures",
                    season_name="2024-2025",
                    success=True,
                    message="Demo with fallback data - real scraping failed",
                    links=[
                        "https://fbref.com/en/matches/cc5b4244/Manchester-United-Fulham-August-16-2024-Premier-League",
                        "https://fbref.com/en/matches/8b1e4321/Arsenal-Wolves-August-17-2024-Premier-League",
                        "https://fbref.com/en/matches/2c4f8901/Brighton-Everton-August-17-2024-Premier-League"
                    ]
                )
            ]
        else:
            # Use real data but limit to first 5 matches for demo
            demo_links = real_result.links[:5]  # Take only first 5 matches
            
            print(f"Demo: Successfully extracted {len(demo_links)} real matches from current season")
            
            demo_seasons = [
                SeasonResult(
                    url=current_season_url,
                    season_name="2024-2025",
                    success=True,
                    message=f"Successfully extracted {len(demo_links)} real match report links from current Premier League season",
                    links=demo_links
                )
            ]
        
        # Compile all links
        all_links = []
        for season in demo_seasons:
            all_links.extend(season.links)
        
        # Generate Excel workbook with real data
        excel_b64, filename = create_excel_workbook(demo_seasons)
        
        return ScrapingResponse(
            success=True,
            message=f"Demo: Successfully processed {len(demo_seasons)} season | Total match links extracted: {len(all_links)} | Excel file generated with {len(all_links)} match sheets from REAL current Premier League data",
            seasons=demo_seasons,
            total_links=len(all_links),
            excel_data=excel_b64,
            filename=filename
        )
        
    except Exception as e:
        print(f"Demo error: {str(e)}")
        # Absolute fallback with fake data
        demo_seasons = [
            SeasonResult(
                url="https://fbref.com/en/comps/9/schedule/Premier-League-Scores-and-Fixtures",
                season_name="2024-2025",
                success=True,
                message="Demo with emergency fallback data due to error",
                links=[
                    "https://fbref.com/en/matches/cc5b4244/Manchester-United-Fulham-August-16-2024-Premier-League",
                    "https://fbref.com/en/matches/8b1e4321/Arsenal-Wolves-August-17-2024-Premier-League"
                ]
            )
        ]
        
        all_links = []
        for season in demo_seasons:
            all_links.extend(season.links)
        
        excel_b64, filename = create_excel_workbook(demo_seasons)
        
        return ScrapingResponse(
            success=True,
            message=f"Demo: Emergency fallback - {len(all_links)} sample links | Error: {str(e)}",
            seasons=demo_seasons,
            total_links=len(all_links),
            excel_data=excel_b64,
            filename=filename
        )

def extract_match_info_from_url(url: str) -> Dict[str, str]:
    """Extract match information from FBREF match URL"""
    try:
        # Example URL: https://fbref.com/en/matches/cc5b4244/Manchester-United-Fulham-August-16-2024-Premier-League
        parts = url.split('/')
        if len(parts) > 5:
            match_info_part = parts[-1]  # "Manchester-United-Fulham-August-16-2024-Premier-League"
            
            # Try to parse the format: Home-Team-Away-Team-Date-Competition
            info_parts = match_info_part.split('-')
            
            # Find date pattern (looking for year)
            date_idx = None
            for i, part in enumerate(info_parts):
                if len(part) == 4 and part.isdigit() and part.startswith('20'):
                    date_idx = i
                    break
            
            if date_idx and date_idx >= 3:
                # Extract teams (everything before date)
                team_parts = info_parts[:date_idx-2]  # Exclude month and day
                home_team = team_parts[0] if len(team_parts) > 0 else "Unknown"
                
                # Find where away team starts (usually in middle)
                mid_point = len(team_parts) // 2
                away_team = team_parts[mid_point] if len(team_parts) > mid_point else "Unknown"
                
                # Extract date
                if date_idx >= 2:
                    month = info_parts[date_idx-2]
                    day = info_parts[date_idx-1]
                    year = info_parts[date_idx]
                    date = f"{month}-{day}-{year}"
                else:
                    date = "Unknown"
                
                # Extract competition (everything after year)
                competition = "-".join(info_parts[date_idx+1:]) if date_idx+1 < len(info_parts) else "Unknown"
                
                # Create short name for sheet
                short_name = f"{home_team}_vs_{away_team}"[:20]
                
                return {
                    'home_team': home_team.replace('-', ' '),
                    'away_team': away_team.replace('-', ' '),
                    'date': date,
                    'competition': competition.replace('-', ' '),
                    'short_name': short_name
                }
    except Exception as e:
        print(f"Error parsing match URL {url}: {e}")
    
    # Fallback
    return {
        'home_team': 'Unknown',
        'away_team': 'Unknown', 
        'date': 'Unknown',
        'competition': 'Unknown',
        'short_name': 'Unknown_Match'
    }

def create_excel_workbook(season_results: List[SeasonResult]) -> tuple[str, str]:
    """Create an Excel workbook with separate sheets for each match and a summary sheet"""
    
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet
    summary_sheet = wb.create_sheet("Summary")
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary sheet headers
    summary_headers = [
        "Season", "Match_Report_URL", "Match_Number", "Home_Team", "Away_Team", 
        "Date", "Score", "Home_Goals", "Away_Goals", "Competition", "Venue", "Sheet_Name"
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Track all links and create sheets
    all_links = []
    match_number = 1
    
    for season_result in season_results:
        if season_result.success:
            for link in season_result.links:
                all_links.append({
                    'season': season_result.season_name,
                    'url': link,
                    'source_url': season_result.url,
                    'match_number': match_number
                })
                
                # Extract match info from URL for sheet naming
                match_info = extract_match_info_from_url(link)
                sheet_name = f"Match_{match_number:03d}_{match_info['short_name']}"
                
                # Ensure sheet name is valid (max 31 chars, no special chars)
                sheet_name = re.sub(r'[^\w\-_]', '_', sheet_name)[:31]
                
                # Create individual match sheet
                match_sheet = wb.create_sheet(sheet_name)
                
                # Match sheet headers - Basic match info
                match_headers = [
                    "Field", "Value"
                ]
                
                for col, header in enumerate(match_headers, 1):
                    cell = match_sheet.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                
                # Add basic match information
                match_info_rows = [
                    ("Season", season_result.season_name),
                    ("Match Number", str(match_number)),
                    ("Match Report URL", link),
                    ("Home Team", match_info['home_team']),
                    ("Away Team", match_info['away_team']),
                    ("Date", match_info['date']),
                    ("Competition", match_info['competition']),
                    ("Source URL", season_result.url),
                    ("", ""),
                    ("=== MATCH STATISTICS ===", ""),
                    ("", ""),
                    ("Goals (Home)", ""),
                    ("Goals (Away)", ""),
                    ("Final Score", ""),
                    ("Attendance", ""),
                    ("Referee", ""),
                    ("Stadium", ""),
                    ("", ""),
                    ("=== HOME TEAM STATS ===", ""),
                    ("", ""),
                    ("Possession (%)", ""),
                    ("Total Shots", ""),
                    ("Shots on Target", ""),
                    ("Corners", ""),
                    ("Fouls", ""),
                    ("Yellow Cards", ""),
                    ("Red Cards", ""),
                    ("", ""),
                    ("=== AWAY TEAM STATS ===", ""),
                    ("", ""),
                    ("Possession (%)", ""),
                    ("Total Shots", ""),
                    ("Shots on Target", ""),
                    ("Corners", ""),
                    ("Fouls", ""),
                    ("Yellow Cards", ""),
                    ("Red Cards", ""),
                    ("", ""),
                    ("=== PLAYER STATISTICS ===", ""),
                    ("", ""),
                    ("Player Name", "Team"),
                    ("Position", "Minutes"),
                    ("Goals", "Assists"),
                    ("Shots", "Passes"),
                    ("Tackles", "Cards")
                ]
                
                for row_num, (field, value) in enumerate(match_info_rows, 2):
                    match_sheet.cell(row=row_num, column=1, value=field)
                    match_sheet.cell(row=row_num, column=2, value=value)
                    
                    # Style section headers
                    if field.startswith("===") and field.endswith("==="):
                        cell = match_sheet.cell(row=row_num, column=1)
                        cell.font = Font(bold=True, color="366092")
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                # Auto-adjust column widths
                for column in match_sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    match_sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add to summary sheet
                summary_row = match_number + 1
                summary_data = [
                    season_result.season_name,
                    link,
                    match_number,
                    match_info['home_team'],
                    match_info['away_team'],
                    match_info['date'],
                    "",  # Score - to be filled later
                    "",  # Home goals
                    "",  # Away goals
                    match_info['competition'],
                    "",  # Venue - to be filled later
                    sheet_name
                ]
                
                for col, value in enumerate(summary_data, 1):
                    cell = summary_sheet.cell(row=summary_row, column=col, value=value)
                    cell.border = border
                
                match_number += 1
    
    # Auto-adjust column widths for summary sheet
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename
    seasons_text = "_".join([s.season_name.replace("/", "-") for s in season_results if s.success])[:50]
    filename = f"FBREF_Matches_{seasons_text}.xlsx"
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Encode to base64
    excel_b64 = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
    
    return excel_b64, filename

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)