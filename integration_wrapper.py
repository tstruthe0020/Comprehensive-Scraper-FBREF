#!/usr/bin/env python3
"""
Integration Wrapper - Bridge between URL compiler app and FBref comprehensive scraper
This file provides the main integration points for the other AI agent to use
"""

import asyncio
import sys
import os
from pathlib import Path
from typing import Dict, Any, Optional
import logging

# Add batch_scraper to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'batch_scraper'))

try:
    from excel_integrator import ExcelIntegrator
    from config import Config
    FBREF_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ FBref scraper not available: {e}")
    FBREF_AVAILABLE = False

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class FBrefIntegration:
    """
    Main integration class for enhancing Excel files with comprehensive FBref data
    
    Usage:
        fbref = FBrefIntegration()
        results = fbref.populate_excel_sync("FBREF_Matches_2024-25.xlsx")
    """
    
    def __init__(self, rate_limit_delay: int = 3, headless: bool = True):
        """
        Initialize FBref integration
        
        Args:
            rate_limit_delay: Seconds between requests (default: 3)
            headless: Run browser in background (default: True)
        """
        if not FBREF_AVAILABLE:
            raise ImportError("FBref scraper dependencies not available")
            
        self.config = Config()
        self.config.RATE_LIMIT_DELAY = rate_limit_delay
        self.config.HEADLESS = headless
        
        # Conservative defaults for production
        self.config.PAGE_TIMEOUT = 30000
        self.config.EXTRACT_PLAYER_STATS = True
        self.config.EXTRACT_MATCH_EVENTS = True
        
    async def populate_excel_file(self, excel_path: str, method: str = "individual") -> Dict[str, Any]:
        """
        Populate an Excel file with comprehensive FBref data
        
        Args:
            excel_path: Path to Excel file with match URLs
            method: "individual" (read from match sheets) or "summary" (read from summary sheet)
            
        Returns:
            dict: Results summary with success/failure counts
        """
        try:
            # Validate file exists
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
            logger.info(f"Starting FBref integration for: {excel_path}")
            
            # Initialize integrator
            integrator = ExcelIntegrator(self.config)
            
            # Choose integration method
            if method == "summary":
                results = await integrator.populate_from_summary_sheet(excel_path)
            else:
                results = await integrator.populate_excel_file(excel_path)
            
            # Format results
            return {
                'success': True,
                'total_matches': results['total_matches'],
                'successful_matches': results['successful_matches'],
                'failed_matches': results['failed_matches'],
                'success_rate': f"{(results['successful_matches']/results['total_matches']*100):.1f}%" if results['total_matches'] > 0 else "0%",
                'errors': results['errors'],
                'output_file': excel_path,
                'method_used': method
            }
            
        except Exception as e:
            logger.error(f"FBref integration failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'output_file': excel_path,
                'method_used': method
            }
    
    def populate_excel_sync(self, excel_path: str, method: str = "individual") -> Dict[str, Any]:
        """
        Synchronous wrapper for async populate function
        
        Args:
            excel_path: Path to Excel file with match URLs
            method: "individual" or "summary"
            
        Returns:
            dict: Results summary
        """
        return asyncio.run(self.populate_excel_file(excel_path, method))
    
    def validate_excel_structure(self, excel_path: str) -> Dict[str, Any]:
        """
        Validate that Excel file has the required structure
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            dict: Validation results
        """
        try:
            import openpyxl
            
            if not os.path.exists(excel_path):
                return {'valid': False, 'error': 'File not found'}
            
            wb = openpyxl.load_workbook(excel_path)
            
            # Check for match sheets
            match_sheets = [s for s in wb.sheetnames if s.startswith("Match_")]
            
            if not match_sheets:
                return {'valid': False, 'error': 'No match sheets found (should start with "Match_")'}
            
            # Check first match sheet structure
            ws = wb[match_sheets[0]]
            
            # Verify URL location (row 4, column 2)
            url = ws.cell(row=4, column=2).value
            if not url:
                return {'valid': False, 'error': 'No URL found in row 4, column 2'}
            
            if not isinstance(url, str) or not url.startswith("https://fbref.com"):
                return {'valid': False, 'error': f'Invalid FBref URL in row 4, column 2: {url}'}
            
            # Check for basic structure (just verify key rows exist, not content)
            required_rows = [
                (5, 2, "Home team"),     # Row 5: Home Team
                (6, 2, "Away team"),     # Row 6: Away Team  
            ]
            
            missing_sections = []
            for row, col, description in required_rows:
                cell_value = ws.cell(row=row, column=col).value
                if not cell_value:
                    missing_sections.append(f"Row {row}, Col {col} ({description})")
            
            # Just check if we have enough rows (our structure should have at least 40 rows)
            max_row = ws.max_row
            if max_row < 30:
                missing_sections.append(f"Insufficient rows: {max_row} (need at least 30)")
            
            if missing_sections:
                return {
                    'valid': False, 
                    'error': f'Missing required sections: {", ".join(missing_sections)}'
                }
            
            return {
                'valid': True,
                'match_sheets_found': len(match_sheets),
                'sample_url': url,
                'structure': 'Compatible'
            }
            
        except Exception as e:
            return {'valid': False, 'error': str(e)}

class FBrefBatchProcessor:
    """
    Batch processor for multiple Excel files or seasons
    """
    
    def __init__(self, rate_limit_delay: int = 4):
        """
        Initialize batch processor with conservative rate limiting
        
        Args:
            rate_limit_delay: Seconds between requests (default: 4 for batch processing)
        """
        self.fbref = FBrefIntegration(rate_limit_delay=rate_limit_delay)
    
    async def process_multiple_files(self, excel_files: list) -> Dict[str, Any]:
        """
        Process multiple Excel files with comprehensive data
        
        Args:
            excel_files: List of Excel file paths
            
        Returns:
            dict: Batch processing results
        """
        results = {
            'total_files': len(excel_files),
            'successful_files': 0,
            'failed_files': 0,
            'file_results': [],
            'overall_matches': 0,
            'overall_successful_matches': 0
        }
        
        for i, excel_file in enumerate(excel_files):
            logger.info(f"Processing file {i+1}/{len(excel_files)}: {excel_file}")
            
            try:
                file_result = await self.fbref.populate_excel_file(excel_file)
                
                if file_result['success']:
                    results['successful_files'] += 1
                    results['overall_matches'] += file_result['total_matches']
                    results['overall_successful_matches'] += file_result['successful_matches']
                else:
                    results['failed_files'] += 1
                
                results['file_results'].append(file_result)
                
            except Exception as e:
                logger.error(f"Error processing {excel_file}: {e}")
                results['failed_files'] += 1
                results['file_results'].append({
                    'success': False,
                    'error': str(e),
                    'output_file': excel_file
                })
        
        # Calculate overall success rate
        if results['overall_matches'] > 0:
            results['overall_success_rate'] = f"{(results['overall_successful_matches']/results['overall_matches']*100):.1f}%"
        else:
            results['overall_success_rate'] = "0%"
        
        return results
    
    def process_multiple_files_sync(self, excel_files: list) -> Dict[str, Any]:
        """Synchronous wrapper for batch processing"""
        return asyncio.run(self.process_multiple_files(excel_files))

# Convenience functions for easy integration
def enhance_excel_with_fbref_data(excel_path: str, rate_limit_delay: int = 3) -> Dict[str, Any]:
    """
    Simple function to enhance an Excel file with FBref data
    
    Args:
        excel_path: Path to Excel file
        rate_limit_delay: Seconds between requests
        
    Returns:
        dict: Enhancement results
    """
    if not FBREF_AVAILABLE:
        return {
            'success': False,
            'error': 'FBref scraper not available',
            'output_file': excel_path
        }
    
    try:
        fbref = FBrefIntegration(rate_limit_delay=rate_limit_delay)
        
        # Check if we're in an async context
        try:
            loop = asyncio.get_running_loop()
            # We're already in an event loop, use a different approach
            import concurrent.futures
            import threading
            
            def run_in_thread():
                # Create new event loop in separate thread
                new_loop = asyncio.new_event_loop()
                asyncio.set_event_loop(new_loop)
                try:
                    return new_loop.run_until_complete(fbref.populate_excel_file(excel_path))
                finally:
                    new_loop.close()
            
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(run_in_thread)
                return future.result(timeout=300)  # 5 minute timeout
                
        except RuntimeError:
            # No event loop running, safe to use asyncio.run
            return fbref.populate_excel_sync(excel_path)
            
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'output_file': excel_path
        }

def validate_excel_for_fbref(excel_path: str) -> Dict[str, Any]:
    """
    Validate Excel file structure for FBref integration.
    
    Our specific structure (different from the GitHub repo):
    - Row 1: Headers ("Field", "Value")
    - Row 2: Season  
    - Row 3: Match Number
    - Row 4: Match Report URL  ← Our actual location
    - Row 5: Home Team
    - Row 6: Away Team
    - Row 7: Date
    - Row 8: Competition
    - Row 9: Source URL
    
    Args:
        excel_path: Path to Excel file
        
    Returns:
        dict: Validation results
    """
    if not FBREF_AVAILABLE:
        return {'valid': False, 'error': 'FBref scraper not available'}
    
    try:
        fbref = FBrefIntegration()
        return fbref.validate_excel_structure(excel_path)
    except Exception as e:
        return {'valid': False, 'error': str(e)}

def check_fbref_availability() -> Dict[str, Any]:
    """
    Check if FBref scraper is available and properly configured
    
    Returns:
        dict: Availability status
    """
    if not FBREF_AVAILABLE:
        return {
            'available': False,
            'error': 'FBref scraper dependencies not installed',
            'required_actions': [
                'cd batch_scraper',
                'pip install -r requirements.txt',
                'playwright install chromium'
            ]
        }
    
    try:
        # Test basic initialization
        config = Config()
        return {
            'available': True,
            'status': 'Ready for integration',
            'rate_limit_delay': config.RATE_LIMIT_DELAY,
            'headless_mode': config.HEADLESS
        }
    except Exception as e:
        return {
            'available': False,
            'error': str(e),
            'required_actions': ['Check batch_scraper configuration']
        }

# Example usage for the other AI agent
if __name__ == "__main__":
    print("🚀 FBref Integration Wrapper")
    print("=" * 40)
    
    # Check availability
    status = check_fbref_availability()
    print(f"FBref Available: {status['available']}")
    
    if status['available']:
        print("✅ Ready for integration!")
        
        # Example usage
        print("\n📋 Example Usage:")
        print("from integration_wrapper import enhance_excel_with_fbref_data")
        print('results = enhance_excel_with_fbref_data("FBREF_Matches_2024-25.xlsx")')
        print("if results['success']:")
        print("    print(f\"Enhanced with {results['successful_matches']} matches!\")")
    else:
        print(f"❌ Setup required: {status['error']}")
        if 'required_actions' in status:
            print("Required actions:")
            for action in status['required_actions']:
                print(f"  - {action}")
