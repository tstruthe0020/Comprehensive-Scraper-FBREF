
import requests
import sys
import json
import time
import base64
import io
import os
from typing import List, Dict, Any
from openpyxl import load_workbook

class FBREFScraperTester:
    def __init__(self, base_url="https://90280e76-97e5-45e5-996a-34bbf3ba566e.preview.emergentagent.com"):
        self.base_url = base_url
        self.tests_run = 0
        self.tests_passed = 0
        self.test_output_dir = "/tmp/fbref_test_output"
        
        # Create output directory if it doesn't exist
        if not os.path.exists(self.test_output_dir):
            os.makedirs(self.test_output_dir)

    def run_test(self, name, method, endpoint, expected_status, data=None):
        """Run a single API test"""
        url = f"{self.base_url}/{endpoint}"
        headers = {'Content-Type': 'application/json'}

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                response = requests.post(url, json=data, headers=headers)

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                print(f"✅ Passed - Status: {response.status_code}")
                return success, response.json() if response.text else {}
            else:
                print(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                print(f"Response: {response.text}")
                return success, {}

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, {}

    def test_health_check(self):
        """Test the health check endpoint"""
        success, response = self.run_test(
            "Health Check",
            "GET",
            "api/health",
            200
        )
        return success

    def save_and_validate_excel(self, excel_data, filename, test_name):
        """Save Excel data to file and validate its structure"""
        if not excel_data:
            print("❌ No Excel data found in response")
            return False
            
        try:
            # Decode base64 data
            excel_bytes = base64.b64decode(excel_data)
            
            # Save to file for inspection
            file_path = os.path.join(self.test_output_dir, f"{test_name}_{filename}")
            with open(file_path, "wb") as f:
                f.write(excel_bytes)
            print(f"✅ Excel file saved to {file_path}")
            
            # Load workbook to validate structure
            wb = load_workbook(io.BytesIO(excel_bytes))
            
            # Check for Summary sheet
            if "Summary" not in wb.sheetnames:
                print("❌ Summary sheet not found in Excel file")
                return False
                
            # Check for match sheets
            match_sheets = [s for s in wb.sheetnames if s != "Summary"]
            if not match_sheets:
                print("❌ No match sheets found in Excel file")
                return False
                
            print(f"✅ Excel file contains Summary sheet + {len(match_sheets)} match sheets")
            
            # Validate Summary sheet structure
            summary_sheet = wb["Summary"]
            expected_headers = [
                "Season", "Match_Report_URL", "Match_Number", "Home_Team", "Away_Team", 
                "Date", "Score", "Home_Goals", "Away_Goals", "Competition", "Venue", "Sheet_Name"
            ]
            
            header_row = [cell.value for cell in summary_sheet[1]]
            missing_headers = [h for h in expected_headers if h not in header_row]
            if missing_headers:
                print(f"❌ Summary sheet missing headers: {missing_headers}")
                return False
                
            # Validate a sample match sheet structure
            sample_sheet = wb[match_sheets[0]]
            
            # Check for Match Report URL in row 3, column 2
            match_url_cell = sample_sheet.cell(row=3, column=2)
            if not match_url_cell.value or not str(match_url_cell.value).startswith("https://fbref.com"):
                print(f"❌ Match Report URL not found in expected location (row 3, column 2)")
                return False
                
            # Check for section headers
            section_headers = ["=== MATCH STATISTICS ===", "=== HOME TEAM STATS ===", "=== AWAY TEAM STATS ===", "=== PLAYER STATISTICS ==="]
            found_headers = []
            
            for row in sample_sheet.iter_rows():
                cell_value = row[0].value
                if cell_value in section_headers:
                    found_headers.append(cell_value)
                    
            missing_sections = [h for h in section_headers if h not in found_headers]
            if missing_sections:
                print(f"❌ Match sheet missing section headers: {missing_sections}")
                return False
                
            print("✅ Excel file structure validated successfully")
            return True
            
        except Exception as e:
            print(f"❌ Excel validation failed: {str(e)}")
            return False

    def test_demo_scrape(self):
        """Test the demo scrape endpoint"""
        success, response = self.run_test(
            "Demo Scrape",
            "POST",
            "api/demo-scrape",
            200,
            data={}
        )
        
        if success:
            print(f"Demo message: {response.get('message', '')}")
            print(f"Total links: {response.get('total_links', 0)}")
            print(f"Number of seasons: {len(response.get('seasons', []))}")
            
            # Check if seasons array exists and has expected structure
            seasons = response.get('seasons', [])
            if not seasons:
                print("❌ No seasons found in response")
                return False
                
            # Check if Excel data exists and validate structure
            excel_validation = self.save_and_validate_excel(
                response.get('excel_data', ''),
                response.get('filename', 'demo_output.xlsx'),
                'demo'
            )
            
            if not excel_validation:
                return False
                
            print("✅ Demo scrape response has expected structure")
            return True
        return False

    def test_single_url_scrape(self, url):
        """Test scraping a single URL"""
        success, response = self.run_test(
            "Single URL Scrape",
            "POST",
            "api/scrape-fbref",
            200,
            data={"urls": [url]}
        )
        
        if success:
            print(f"Scrape message: {response.get('message', '')}")
            print(f"Total links: {response.get('total_links', 0)}")
            
            # Check if seasons array exists and has expected structure
            seasons = response.get('seasons', [])
            if not seasons:
                print("❌ No seasons found in response")
                return False
                
            # Check if the season was successfully scraped
            if not seasons[0].get('success'):
                print(f"❌ Scraping failed: {seasons[0].get('message', '')}")
                return False
                
            # Check if links were extracted
            links = seasons[0].get('links', [])
            if not links:
                print("❌ No links extracted")
                return False
                
            print(f"✅ Successfully extracted {len(links)} links")
            
            # Check if Excel data exists and validate structure
            excel_validation = self.save_and_validate_excel(
                response.get('excel_data', ''),
                response.get('filename', 'single_url_output.xlsx'),
                'single'
            )
            
            if not excel_validation:
                return False
                
            print("✅ Single URL scrape response has expected structure")
            return True
        return False

    def test_multi_url_scrape(self, urls):
        """Test scraping multiple URLs"""
        success, response = self.run_test(
            "Multi URL Scrape",
            "POST",
            "api/scrape-fbref",
            200,
            data={"urls": urls}
        )
        
        if success:
            print(f"Scrape message: {response.get('message', '')}")
            print(f"Total links: {response.get('total_links', 0)}")
            
            # Check if seasons array exists and has expected structure
            seasons = response.get('seasons', [])
            if not seasons:
                print("❌ No seasons found in response")
                return False
                
            # Check if at least one season was successfully scraped
            successful_seasons = [s for s in seasons if s.get('success')]
            if not successful_seasons:
                print("❌ No seasons were successfully scraped")
                return False
                
            print(f"✅ Successfully scraped {len(successful_seasons)}/{len(urls)} seasons")
            
            # Check if Excel data exists and validate structure
            excel_validation = self.save_and_validate_excel(
                response.get('excel_data', ''),
                response.get('filename', 'multi_url_output.xlsx'),
                'multi'
            )
            
            if not excel_validation:
                return False
                
            print("✅ Multi URL scrape response has expected structure")
            return True
        return False

    def test_invalid_url(self):
        """Test scraping with an invalid URL"""
        success, response = self.run_test(
            "Invalid URL Scrape",
            "POST",
            "api/scrape-fbref",
            400,  # Expecting a 400 Bad Request
            data={"urls": ["https://example.com"]}
        )
        
        # For this test, we expect a failure with a 400 status code
        if not success and response:
            print("✅ Invalid URL correctly rejected")
            return True
        return False

    def test_mixed_urls(self, valid_url, invalid_url):
        """Test scraping with a mix of valid and invalid URLs"""
        success, response = self.run_test(
            "Mixed URLs Scrape",
            "POST",
            "api/scrape-fbref",
            200,  # Expecting a 200 OK with partial success
            data={"urls": [valid_url, invalid_url]}
        )
        
        if success:
            # Check if seasons array exists
            seasons = response.get('seasons', [])
            if not seasons or len(seasons) != 2:
                print("❌ Expected 2 seasons in response")
                return False
                
            # Check if one season succeeded and one failed
            successful = [s for s in seasons if s.get('success')]
            failed = [s for s in seasons if not s.get('success')]
            
            if len(successful) != 1 or len(failed) != 1:
                print(f"❌ Expected 1 success and 1 failure, got {len(successful)} successes and {len(failed)} failures")
                return False
                
            print("✅ Mixed URLs handled correctly")
            return True
        return False
        
    def test_real_premier_league_seasons(self):
        """Test with real Premier League seasons 2022-2023 and 2023-2024"""
        urls = [
            "https://fbref.com/en/comps/9/2022-2023/schedule/2022-2023-Premier-League-Scores-and-Fixtures",
            "https://fbref.com/en/comps/9/2023-2024/schedule/2023-2024-Premier-League-Scores-and-Fixtures"
        ]
        
        print("\n🔍 Testing Real Premier League Seasons (2022-2023 and 2023-2024)...")
        print("⚠️ This test may take several minutes to complete due to the large dataset")
        
        success, response = self.run_test(
            "Premier League Seasons Scrape",
            "POST",
            "api/scrape-fbref",
            200,
            data={"urls": urls}
        )
        
        if success:
            print(f"Scrape message: {response.get('message', '')}")
            total_links = response.get('total_links', 0)
            print(f"Total links: {total_links}")
            
            # Check if we have a large number of matches (expecting around 760)
            if total_links < 700:
                print(f"❌ Expected around 760 matches, but only found {total_links}")
                return False
                
            print(f"✅ Successfully extracted {total_links} match links")
            
            # Check if Excel data exists and validate structure
            excel_validation = self.save_and_validate_excel(
                response.get('excel_data', ''),
                response.get('filename', 'premier_league_seasons.xlsx'),
                'premier_league'
            )
            
            if not excel_validation:
                return False
                
            print("✅ Premier League seasons scrape successful")
            return True
        return False

def main():
    # Setup
    tester = FBREFScraperTester()
    
    # Valid FBREF URLs for testing
    premier_league_url = "https://fbref.com/en/comps/9/schedule/Premier-League-Scores-and-Fixtures"
    premier_league_2023_url = "https://fbref.com/en/comps/9/2023-2024/schedule/2023-2024-Premier-League-Scores-and-Fixtures"
    invalid_url = "https://fbref.com/invalid/url"
    
    # Run tests
    print("\n===== FBREF Multi-Season Scraper API Tests =====\n")
    
    # Test 1: Health Check
    health_check_success = tester.test_health_check()
    
    # Test 2: Demo Scrape
    demo_scrape_success = tester.test_demo_scrape()
    
    # Test 3: Single URL Scrape
    print("\nTesting with a single URL. This may take some time...")
    single_url_success = tester.test_single_url_scrape(premier_league_url)
    
    # Test 4: Multi URL Scrape
    print("\nTesting with multiple URLs. This may take some time...")
    multi_url_success = tester.test_multi_url_scrape([premier_league_url, premier_league_2023_url])
    
    # Test 5: Invalid URL
    invalid_url_success = tester.test_invalid_url()
    
    # Test 6: Mixed URLs
    mixed_urls_success = tester.test_mixed_urls(premier_league_url, invalid_url)
    
    # Print results
    print(f"\n📊 Tests passed: {tester.tests_passed}/{tester.tests_run}")
    
    # Summary
    print("\n===== Test Summary =====")
    print(f"Health Check: {'✅' if health_check_success else '❌'}")
    print(f"Demo Scrape: {'✅' if demo_scrape_success else '❌'}")
    print(f"Single URL Scrape: {'✅' if single_url_success else '❌'}")
    print(f"Multi URL Scrape: {'✅' if multi_url_success else '❌'}")
    print(f"Invalid URL Handling: {'✅' if invalid_url_success else '❌'}")
    print(f"Mixed URLs Handling: {'✅' if mixed_urls_success else '❌'}")
    
    return 0 if tester.tests_passed == tester.tests_run else 1

if __name__ == "__main__":
    sys.exit(main())
      